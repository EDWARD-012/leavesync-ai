"""
AI-powered holiday parser with structured fallbacks.

Usage:
    from core.utils.holiday_parser import parse_holiday_file
    holidays = parse_holiday_file(uploaded_file)  # -> List[ (name: str, date: datetime.date) ]

This module:
 - Tries local parsing (openpyxl, pdfplumber, pytesseract)
 - If results are insufficient, calls Gemini (text model) to extract holidays as structured JSON
 - Returns list of tuples (holiday_name, date)

Environment:
 - Set GEMINI_API_KEY and optionally GEMINI_MODEL in Django settings or .env
    GEMINI_API_KEY=...
    GEMINI_MODEL=gemini-1.5-preview (or your preferred model)
"""

import json
import os
import re
from datetime import datetime, date
from typing import List, Tuple, Optional
from django.conf import settings

# Optional libs
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import pdfplumber

    PDFPLUMBER_AVAILABLE = True
except Exception:
    PDFPLUMBER_AVAILABLE = False

try:
    from PIL import Image
    import pytesseract

    PYTESSERACT_AVAILABLE = True
except Exception:
    PYTESSERACT_AVAILABLE = False

# Gemini SDK (text fallback). Make sure google-generativeai is installed.
GEMINI_AVAILABLE = False
try:
    import google.generativeai as genai

    GEMINI_AVAILABLE = True
    GEMINI_API_KEY = getattr(settings, "GEMINI_API_KEY", None) or os.getenv("GEMINI_API_KEY")
    GEMINI_MODEL = getattr(settings, "GEMINI_MODEL", "gemini-1.5-flash")
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
    else:
        GEMINI_AVAILABLE = False
except Exception:
    GEMINI_AVAILABLE = False


DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%B %d, %Y",
    "%b %d, %Y",
    "%d %B %Y",
    "%d %b %Y",
]

# Regex patterns to find date-like tokens in lines
DATE_REGEXES = [
    r"\b(\d{4}[/-]\d{1,2}[/-]\d{1,2})\b",  # YYYY-MM-DD or YYYY/MM/DD
    r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{4})\b",  # DD/MM/YYYY or MM/DD/YYYY
    r"\b([A-Za-z]+ \d{1,2}, \d{4})\b",     # Monthname DD, YYYY
    r"\b(\d{1,2} [A-Za-z]+ \d{4})\b",      # DD Monthname YYYY
]


def parse_holiday_file(file) -> List[Tuple[str, date]]:
    """
    Main entry. Accepts Django UploadedFile or a file-like object with .name and file content.
    Returns list of (holiday_name, date) tuples (date is datetime.date).
    """
    filename = getattr(file, "name", "") or ""
    filename = filename.lower()

    candidates: List[Tuple[str, date]] = []

    # 1) Excel
    if OPENPYXL_AVAILABLE and (filename.endswith(".xlsx") or filename.endswith(".xls")):
        candidates = _parse_xlsx(file)
        if _is_good_result(candidates):
            return candidates

    # 2) PDF text extraction
    if PDFPLUMBER_AVAILABLE and filename.endswith(".pdf"):
        candidates = _parse_pdf_text(file)
        if _is_good_result(candidates):
            return candidates

    # 3) Image OCR (jpg/png)
    if PYTESSERACT_AVAILABLE and any(filename.endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".tiff")):
        candidates = _parse_image(file)
        if _is_good_result(candidates):
            return candidates

    # 4) Try simple regex on file bytes/text as a last local attempt
    try:
        file.seek(0)
        raw = file.read()
        if isinstance(raw, bytes):
            raw_text = raw.decode("utf-8", "ignore")
        else:
            raw_text = str(raw)
        candidates = _parse_text_with_regex(raw_text)
        if _is_good_result(candidates):
            return candidates
    except Exception:
        pass

    # 5) Fallback to Gemini if available
    if GEMINI_AVAILABLE:
        try:
            file.seek(0)
            content_bytes = file.read()
            # If file is bytes, try decode; if not, convert to string summary
            if isinstance(content_bytes, bytes):
                summary_text = _extract_text_preview(content_bytes, filename)
            else:
                summary_text = str(content_bytes)[:15000]

            return _call_gemini_for_holidays(summary_text)
        except Exception:
            return []

    # Nothing worked
    return []


# ------------------------------
# Local parsers
# ------------------------------
def _parse_xlsx(file) -> List[Tuple[str, date]]:
    try:
        file.seek(0)
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        holidays = []

        # Heuristic: try first 20 rows to detect date/name columns
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            if not any(row):
                continue
            date_val = None
            name_val = None
            for cell in row:
                if cell is None:
                    continue
                # If cell is actual date object
                if isinstance(cell, datetime):
                    date_val = cell.date()
                # If cell looks like a date string
                elif isinstance(cell, str):
                    parsed = _try_parse_date(cell)
                    if parsed and not date_val:
                        date_val = parsed
                    elif len(cell.strip()) > 2 and not name_val:
                        name_val = cell.strip()
                # Numbers that could be Excel serial dates - handled by openpyxl as datetime normally
            if date_val and not name_val:
                # try to find name in adjacent cells (same row)
                for cell in row:
                    if isinstance(cell, str) and len(cell.strip()) > 2 and not _looks_like_date(cell):
                        name_val = cell.strip()
                        break
            if date_val and name_val:
                holidays.append((name_val, date_val))
        return holidays
    except Exception:
        return []


def _parse_pdf_text(file) -> List[Tuple[str, date]]:
    try:
        file.seek(0)
        text = ""
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
        return _parse_text_with_regex(text)
    except Exception:
        return []


def _parse_image(file) -> List[Tuple[str, date]]:
    try:
        file.seek(0)
        img = Image.open(file)
        text = pytesseract.image_to_string(img)
        return _parse_text_with_regex(text)
    except Exception:
        return []


# ------------------------------
# Text parsing + regex
# ------------------------------
def _parse_text_with_regex(text: str) -> List[Tuple[str, date]]:
    results = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines:
        # attempt to find a date token in the line
        date_token = None
        for regex in DATE_REGEXES:
            m = re.search(regex, line)
            if m:
                date_token = m.group(1)
                break
        if not date_token:
            continue
        parsed = _try_parse_date(date_token)
        if not parsed:
            continue
        # Remove date token from line to extract name
        name = line.replace(date_token, "").strip(" -,:;.()")
        # If after removal name is empty, try previous/next line heuristics
        if not name and len(lines) >= 2:
            # try neighbor lines
            idx = lines.index(line)
            if idx > 0:
                name = lines[idx - 1]
            elif idx + 1 < len(lines):
                name = lines[idx + 1]
        name = _clean_name(name)
        if name:
            results.append((name, parsed))
    return results


def _try_parse_date(date_str: str) -> Optional[date]:
    date_str = date_str.strip()
    # Normalize slashes
    date_str = date_str.replace(".", "/")
    # Try multiple formats
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(date_str, fmt).date()
        except Exception:
            continue
    # Try smart parsing for ambiguous numeric dates (fallback)
    # If string like 01/26/2025 -> try month/day/year
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", date_str)
    if m:
        a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # Try mm/dd/yyyy then dd/mm/yyyy
        try:
            return date(c, a, b)
        except Exception:
            try:
                return date(c, b, a)
            except Exception:
                return None
    return None


def _looks_like_date(s: str) -> bool:
    for regex in DATE_REGEXES:
        if re.search(regex, s):
            return True
    return False


def _clean_name(s: str) -> str:
    s = re.sub(r"\s{2,}", " ", s)
    s = s.strip(" -,:;.()")
    # sometimes names include numbers or stray tokens - keep readable text only
    s = re.sub(r"[^A-Za-z0-9 \-&()']", "", s).strip()
    return s if len(s) > 1 else ""


def _is_good_result(candidates: List[Tuple[str, date]]) -> bool:
    # Heuristic: at least 2 holidays found OR 1 holiday with a full date in current/next year
    if not candidates:
        return False
    if len(candidates) >= 2:
        return True
    # single candidate: ensure date is sensible (year within next 2 years)
    single = candidates[0]
    if single and isinstance(single[1], date):
        yr = single[1].year
        now = datetime.now().year
        if now <= yr <= now + 2:
            return True
    return False


# ------------------------------
# Gemini fallback
# ------------------------------
def _call_gemini_for_holidays(text_preview: str) -> List[Tuple[str, date]]:
    """
    Send the extracted text preview to Gemini (text) and ask for structured JSON.
    Returns list of (name, date) tuples.
    """
    try:
        prompt = (
            "You are an expert data extractor. Given the following text that contains a company's "
            "holiday list (may be messy, multi-column, include dates in different formats), "
            "extract all holidays and return JSON ONLY in this exact structure:\n\n"
            '[{"name": "Holiday Name", "date": "YYYY-MM-DD"}, ...]\n\n'
            "Text:\n\n"
            f"{text_preview}\n\n"
            "Important:\n"
            "- Output valid JSON (array). Use ISO date format YYYY-MM-DD. "
            "If a holiday has only day/month and not year, assume nearest upcoming year. "
            "- If you find duplicates, return them once.\n"
        )

        model = genai.GenerativeModel(GEMINI_MODEL)
        response = model.generate_content(prompt)

        if not response or not response.text:
            return []

        text = response.text.strip()
        # Try to extract JSON block from response
        json_text = _extract_json_block(text)
        if not json_text:
            json_text = text  # maybe the model returned pure JSON

        parsed = json.loads(json_text)
        results = []
        now_year = datetime.now().year
        for item in parsed:
            nm = item.get("name") or item.get("holiday") or ""
            dt = item.get("date") or item.get("day") or ""
            if not nm or not dt:
                continue
            # normalize date - if missing year try attach nearest year
            try:
                parsed_date = _try_parse_date(dt)
                if parsed_date is None:
                    # try parsing without year -> attach nearest year
                    m = re.match(r"^(\d{1,2})[/-](\d{1,2})$", dt)
                    if m:
                        mm = int(m.group(1)); dd = int(m.group(2))
                        # choose this year or next
                        try_date = date(now_year, mm, dd)
                        if try_date < date.today():
                            try_date = date(now_year + 1, mm, dd)
                        parsed_date = try_date
                if parsed_date:
                    results.append((_clean_name(nm), parsed_date))
            except Exception:
                continue

        # dedupe by (name,date)
        uniq = []
        seen = set()
        for name, dt in results:
            key = (name.lower(), dt.isoformat())
            if key in seen:
                continue
            seen.add(key)
            uniq.append((name, dt))
        return uniq
    except Exception:
        return []


def _extract_json_block(s: str) -> Optional[str]:
    """
    Try to safely extract a JSON array/object block from the model response.
    """
    # look for first '[' and last ']'
    start = s.find("[")
    end = s.rfind("]")
    if start != -1 and end != -1 and end > start:
        return s[start : end + 1]
    # fallback to curly
    start = s.find("{")
    end = s.rfind("}")
    if start != -1 and end != -1 and end > start:
        return s[start : end + 1]
    return None


# ------------------------------
# Helpers
# ------------------------------
def _extract_text_preview(content_bytes: bytes, filename: str) -> str:
    """
    Create a reasonably sized text preview to send to Gemini:
    - If text-like, decode first N chars
    - If binary (image/pdf), return short metadata + first 6000 bytes decoded
    """
    try:
        # try decode
        text = content_bytes.decode("utf-8")
        return text[:12000]
    except Exception:
        # binary file -> return filename and partial bytes (base64 avoided)
        snippet = ""
        try:
            snippet = content_bytes[:12000].hex()[:4000]
        except Exception:
            snippet = filename or "file"
        return f"FILE: {filename}\n\n[Binary file preview not decodable]\n\n{snippet}"


# ------------------------------
# End of module
# ------------------------------
